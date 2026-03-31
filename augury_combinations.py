from augury_knowledge import ZiWeiKnowledgeBase, Brightness
import pandas as pd
import re
from typing import List
import openpyxl
class ZiWeiCombinationProcessor:
    """紫微斗数组合处理器"""
    
    def __init__(self, knowledge_base: ZiWeiKnowledgeBase):
        self.kb = knowledge_base
        self.brightness_cache = {}  # 亮度缓存
        
    def process_combinations_file(self, input_file: str, output_file: str):
        """处理组合文件"""
        try:
            # 读取组合文件
            df = pd.read_excel(input_file)
            
            # 处理每个组合
            processed_data = []
            
            for index, row in df.iterrows():
                combination = row['星曜组合']
                stars = self._parse_star_combination(combination)
                
                # 为每个场景生成象义
                scenarios = self.kb.get_all_scenarios()
                
                for scenario in scenarios:
                    meaning = self.generate_combination_meaning(stars, scenario)
                    
                    processed_data.append({
                        '星曜组合': combination,
                        '场景维度': scenario,
                        '象义组合': meaning,
                        '星曜列表': ','.join(stars)
                    })
            
            # 创建结果DataFrame
            result_df = pd.DataFrame(processed_data)
            
            # 保存结果
            self._save_formatted_result(result_df, output_file)
            
            print(f"处理完成！共生成 {len(processed_data)} 条象义组合")
            return result_df
            
        except Exception as e:
            print(f"处理组合文件错误: {e}")
            return None
    
    def _parse_star_combination(self, combination: str) -> List[str]:
        """解析星曜组合字符串"""
        if pd.isna(combination):
            return []
        
        # 支持多种分隔符：逗号、空格、顿号等
        stars = re.split(r'[,，\s、]+', combination.strip())
        return [star for star in stars if star]
    
    def generate_combination_meaning(self, stars: List[str], scenario: str) -> str:
        """生成星曜组合的象义"""
        if not stars:
            return "无星曜组合"
        
        meanings = []
        
        # 1. 首先检查是否有特殊组合效应
        combination_effect = self.kb.get_combination_effect(stars, scenario)
        if combination_effect:
            meanings.append(f"【组合特效】{combination_effect}")
        
        # 2. 为每个星曜生成象义（考虑亮度影响）
        for star in stars:
            star_meanings = []
            
            # 考虑不同亮度的影响
            for brightness in [Brightness.TEMPLE, Brightness.WEAK]:
                meaning = self.kb.get_star_meaning(star, scenario, brightness)
                if meaning and "暂无" not in meaning:
                    star_meanings.append(f"{brightness.value}: {meaning}")
            
            if star_meanings:
                meanings.append(f"{star}→{'；'.join(star_meanings)}")
        
        # 3. 如果没有找到具体象义，提供通用解释
        if not meanings:
            star_names = '、'.join(stars)
            meanings.append(f"{star_names}在{scenario}领域产生交互影响")
        
        return '\n'.join(meanings)
    
    def _save_formatted_result(self, df: pd.DataFrame, output_file: str):
        """保存格式化的结果"""
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='紫占组合象义', index=False)
            
            # 获取工作表进行格式调整
            worksheet = writer.sheets['紫占组合象义']
            
            # 设置列宽和自动换行
            for column in worksheet.columns:
                column_letter = column[0].column_letter
                max_length = 0
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # 设置自动换行
                for cell in column:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
            
            # 合并星曜组合列（按需求）
            self._merge_combination_cells(worksheet, df)
    
    def _merge_combination_cells(self, worksheet, df):
        """合并星曜组合单元格"""
        from openpyxl.utils import get_column_letter
        
        # 计算需要合并的行范围
        current_combination = None
        start_row = 2  # 从数据行开始（第1行是标题）
        
        for i, combination in enumerate(df['星曜组合']):
            if combination != current_combination:
                if current_combination is not None:
                    # 合并上一个组合的单元格
                    end_row = start_row + count - 1
                    if count > 1:  # 只有多行才需要合并
                        worksheet.merge_cells(f'A{start_row}:A{end_row}')
                
                current_combination = combination
                start_row = 2 + i
                count = 1
            else:
                count += 1
        
        # 合并最后一个组合
        if count > 1:
            end_row = start_row + count - 1
            worksheet.merge_cells(f'A{start_row}:A{end_row}')